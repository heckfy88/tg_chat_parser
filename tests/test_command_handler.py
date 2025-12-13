import json
import pytest
from datetime import datetime
from io import BytesIO
from unittest.mock import MagicMock, AsyncMock, patch
from openpyxl import load_workbook

from bot.handlers.command_handler import BotCommandHandler, generate_excel, extract_text


class TestExtractText:
    """Тесты для функции extract_text"""

    def test_extract_text_from_string(self):
        """Тест извлечения текста из строки"""
        result = extract_text("Simple text")
        assert result == "Simple text"

    def test_extract_text_from_list_of_strings(self):
        """Тест извлечения текста из списка строк"""
        result = extract_text(["Hello", " ", "World"])
        assert result == "Hello World"

    def test_extract_text_from_list_with_dicts(self):
        """Тест извлечения текста из списка с словарями"""
        text_field = [
            "Hello ",
            {"text": "World"},
            "!"
        ]
        result = extract_text(text_field)
        assert result == "Hello World!"

    def test_extract_text_from_mixed_list(self):
        """Тест извлечения текста из смешанного списка"""
        text_field = [
            "Start ",
            {"text": "middle"},
            " end"
        ]
        result = extract_text(text_field)
        assert result == "Start middle end"

    def test_extract_text_from_empty_string(self):
        """Тест извлечения текста из пустой строки"""
        result = extract_text("")
        assert result == ""

    def test_extract_text_from_empty_list(self):
        """Тест извлечения текста из пустого списка"""
        result = extract_text([])
        assert result == ""

    def test_extract_text_from_invalid_type(self):
        """Тест извлечения текста из невалидного типа"""
        result = extract_text(123)
        assert result == ""

    def test_extract_text_from_dict_without_text_key(self):
        """Тест извлечения текста из словаря без ключа text"""
        text_field = [{"other_key": "value"}]
        result = extract_text(text_field)
        assert result == ""


class TestGenerateExcel:
    """Тесты для функции generate_excel"""

    def test_generate_excel_basic(self):
        """Тест генерации Excel файла с базовыми данными"""
        participants = {
            "user123": {"username": "john_doe"},
            "user456": {"username": "jane_smith"}
        }
        output = BytesIO()
        generate_excel(participants, output)
        
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        
        assert ws.title == "Participants"
        assert ws.max_row == 3
        
        headers = [cell.value for cell in ws[1]]
        assert headers == ["Дата экспорта", "UserID", "Nickname"]
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2
        assert ("user123", "john_doe") in [(row[1], row[2]) for row in rows]
        assert ("user456", "jane_smith") in [(row[1], row[2]) for row in rows]

    def test_generate_excel_empty_participants(self):
        """Тест генерации Excel файла с пустым списком участников"""
        participants = {}
        output = BytesIO()
        generate_excel(participants, output)
        
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        
        assert ws.max_row == 1

    def test_generate_excel_with_missing_username(self):
        """Тест генерации Excel файла с отсутствующим username"""
        participants = {
            "user123": {},
            "user456": {"username": "jane_smith"}
        }
        output = BytesIO()
        generate_excel(participants, output)
        
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        usernames = [row[2] for row in rows]
        assert "" in usernames or None in usernames
        assert "jane_smith" in usernames


class TestBotCommandHandler:
    """Тесты для класса BotCommandHandler"""

    @pytest.fixture
    def handler(self, mock_env_vars):
        """Фикстура для создания экземпляра BotCommandHandler"""
        return BotCommandHandler()

    def test_init_with_valid_threshold(self, mock_env_vars):
        """Тест инициализации с валидным порогом"""
        handler = BotCommandHandler()
        assert handler._excel_user_threshold == 10

    def test_init_with_negative_threshold(self, monkeypatch):
        """Тест инициализации с отрицательным порогом"""
        monkeypatch.setenv('EXCEL_USER_THRESHOLD', '-1')
        import importlib
        import bot.handlers.command_handler
        importlib.reload(bot.handlers.command_handler)
        from bot.handlers.command_handler import BotCommandHandler
        
        with pytest.raises(Exception, match="Excel user threshold cannot be negative"):
            BotCommandHandler()

    @pytest.mark.asyncio
    async def test_start_command(self, handler, mock_update, mock_context):
        """Тест команды /start"""
        await handler.start(mock_update, mock_context)
        
        assert mock_context.user_data["files"] == []
        mock_update.message.reply_text.assert_called_once()
        call_args = mock_update.message.reply_text.call_args[0][0]
        assert "Hi! I can analyze exported Telegram chat data" in call_args
        assert "Export your Telegram chat" in call_args

    @pytest.mark.asyncio
    async def test_extract_participants_from_files_basic(self, handler, mock_update, mock_context, sample_telegram_json):
        """Тест извлечения участников из файлов"""
        mock_document = MagicMock()
        mock_file = AsyncMock()
        mock_file.download_as_bytearray = AsyncMock(
            return_value=bytearray(json.dumps(sample_telegram_json).encode('utf-8'))
        )
        mock_document.get_file = AsyncMock(return_value=mock_file)
        
        mock_context.user_data["files"] = [mock_document]
        
        participants_by_id, participants_by_username = await handler.extract_participants_from_files(
            mock_update, mock_context
        )
        
        assert "user123" in participants_by_id
        assert participants_by_id["user123"]["username"] == "john_doe"
        assert "user456" in participants_by_id
        assert participants_by_id["user456"]["username"] == "jane_smith"
        assert "user789" not in participants_by_id
        assert "@alice" in participants_by_username
        assert "@bob" in participants_by_username

    @pytest.mark.asyncio
    async def test_extract_participants_from_files_with_username_regex(self, handler, mock_update, mock_context):
        """Тест извлечения упоминаний через регулярное выражение"""
        json_data = {
            "messages": [
                {
                    "from_id": "user123",
                    "from": "john_doe",
                    "text": "Hello @alice and @bob_test123!",
                    "text_entities": []
                }
            ]
        }
        
        mock_document = MagicMock()
        mock_file = AsyncMock()
        mock_file.download_as_bytearray = AsyncMock(
            return_value=bytearray(json.dumps(json_data).encode('utf-8'))
        )
        mock_document.get_file = AsyncMock(return_value=mock_file)
        
        mock_context.user_data["files"] = [mock_document]
        
        participants_by_id, participants_by_username = await handler.extract_participants_from_files(
            mock_update, mock_context
        )
        
        assert "@alice" in participants_by_username
        assert "@bob_test123" in participants_by_username

    @pytest.mark.asyncio
    async def test_extract_participants_from_files_invalid_json(self, handler, mock_update, mock_context):
        """Тест обработки невалидного JSON"""
        mock_document = MagicMock()
        mock_file = AsyncMock()
        mock_file.download_as_bytearray = AsyncMock(
            return_value=bytearray(b"invalid json {")
        )
        mock_document.get_file = AsyncMock(return_value=mock_file)
        
        mock_context.user_data["files"] = [mock_document]
        
        result = await handler.extract_participants_from_files(
            mock_update, mock_context
        )
        
        assert result is None

    @pytest.mark.asyncio
    async def test_extract_participants_from_files_empty_messages(self, handler, mock_update, mock_context):
        """Тест обработки файла с пустым списком сообщений"""
        json_data = {"messages": []}
        
        mock_document = MagicMock()
        mock_file = AsyncMock()
        mock_file.download_as_bytearray = AsyncMock(
            return_value=bytearray(json.dumps(json_data).encode('utf-8'))
        )
        mock_document.get_file = AsyncMock(return_value=mock_file)
        
        mock_context.user_data["files"] = [mock_document]
        
        participants_by_id, participants_by_username = await handler.extract_participants_from_files(
            mock_update, mock_context
        )
        
        assert participants_by_id == {}
        assert participants_by_username == {}

    @pytest.mark.asyncio
    async def test_extract_participants_from_files_missing_from_id(self, handler, mock_update, mock_context):
        """Тест обработки сообщений без from_id"""
        json_data = {
            "messages": [
                {
                    "from": "john_doe",
                    "text": "Hello",
                    "text_entities": []
                }
            ]
        }
        
        mock_document = MagicMock()
        mock_file = AsyncMock()
        mock_file.download_as_bytearray = AsyncMock(
            return_value=bytearray(json.dumps(json_data).encode('utf-8'))
        )
        mock_document.get_file = AsyncMock(return_value=mock_file)
        
        mock_context.user_data["files"] = [mock_document]
        
        participants_by_id, participants_by_username = await handler.extract_participants_from_files(
            mock_update, mock_context
        )
        
        assert participants_by_id == {}

    @pytest.mark.asyncio
    async def test_process_command_text_output(self, handler, mock_update, mock_context, sample_telegram_json):
        """Тест команды /process с текстовым выводом (мало участников)"""
        handler._excel_user_threshold = 10
        
        mock_document = MagicMock()
        mock_file = AsyncMock()
        mock_file.download_as_bytearray = AsyncMock(
            return_value=bytearray(json.dumps(sample_telegram_json).encode('utf-8'))
        )
        mock_document.get_file = AsyncMock(return_value=mock_file)
        
        mock_context.user_data["files"] = [mock_document]
        
        result = await handler.process(mock_update, mock_context)
        
        mock_update.message.reply_text.assert_called_once()
        call_args = mock_update.message.reply_text.call_args[0][0]
        assert "Результаты анализа файлов" in call_args
        assert "Участники чата" in call_args
        assert "Упоминания" in call_args
        assert isinstance(result, list)

    @pytest.mark.asyncio
    async def test_process_command_excel_output(self, handler, mock_update, mock_context, sample_telegram_json):
        """Тест команды /process с выводом в Excel (много участников)"""
        handler._excel_user_threshold = 1
        
        large_json = {
            "messages": [
                {
                    "from_id": f"user{i}",
                    "from": f"user_{i}",
                    "text": "Hello",
                    "text_entities": []
                }
                for i in range(5)
            ]
        }
        
        mock_document = MagicMock()
        mock_file = AsyncMock()
        mock_file.download_as_bytearray = AsyncMock(
            return_value=bytearray(json.dumps(large_json).encode('utf-8'))
        )
        mock_document.get_file = AsyncMock(return_value=mock_file)
        
        mock_context.user_data["files"] = [mock_document]
        
        result = await handler.process(mock_update, mock_context)
        
        mock_update.message.reply_document.assert_called_once()
        call_kwargs = mock_update.message.reply_document.call_args[1]
        assert "document" in call_kwargs
        assert "filename" in call_kwargs
        assert call_kwargs["filename"].endswith(".xlsx")
        assert result == "Файл отправлен"

    @pytest.mark.asyncio
    async def test_process_command_no_participants(self, handler, mock_update, mock_context):
        """Тест команды /process без участников"""
        json_data = {"messages": []}
        
        mock_document = MagicMock()
        mock_file = AsyncMock()
        mock_file.download_as_bytearray = AsyncMock(
            return_value=bytearray(json.dumps(json_data).encode('utf-8'))
        )
        mock_document.get_file = AsyncMock(return_value=mock_file)
        
        mock_context.user_data["files"] = [mock_document]
        
        result = await handler.process(mock_update, mock_context)
        
        mock_update.message.reply_text.assert_called_once()
        call_args = mock_update.message.reply_text.call_args[0][0]
        assert "Нет участников" in call_args

    @pytest.mark.asyncio
    async def test_extract_participants_multiple_files(self, handler, mock_update, mock_context):
        """Тест извлечения участников из нескольких файлов"""
        json_data_1 = {
            "messages": [
                {
                    "from_id": "user1",
                    "from": "user_one",
                    "text": "Hello",
                    "text_entities": []
                }
            ]
        }
        
        json_data_2 = {
            "messages": [
                {
                    "from_id": "user2",
                    "from": "user_two",
                    "text": "Hi",
                    "text_entities": []
                }
            ]
        }
        
        mock_doc1 = MagicMock()
        mock_file1 = AsyncMock()
        mock_file1.download_as_bytearray = AsyncMock(
            return_value=bytearray(json.dumps(json_data_1).encode('utf-8'))
        )
        mock_doc1.get_file = AsyncMock(return_value=mock_file1)
        
        mock_doc2 = MagicMock()
        mock_file2 = AsyncMock()
        mock_file2.download_as_bytearray = AsyncMock(
            return_value=bytearray(json.dumps(json_data_2).encode('utf-8'))
        )
        mock_doc2.get_file = AsyncMock(return_value=mock_file2)
        
        mock_context.user_data["files"] = [mock_doc1, mock_doc2]
        
        participants_by_id, participants_by_username = await handler.extract_participants_from_files(
            mock_update, mock_context
        )
        
        assert "user1" in participants_by_id
        assert participants_by_id["user1"]["username"] == "user_one"
        assert "user2" not in participants_by_id

